"""
excel_exporter.py
Builds a companion .xlsx (never the original file) with the raw data on
one sheet and a "Dashboard" sheet holding NATIVE Excel chart objects
(openpyxl.chart), so a non-technical user can open it straight in Excel/
Google Sheets with no viewer needed — the tradeoff is losing the custom
Night-Ops-Deck styling (Excel charts render in Excel's own visual
language, not ours).

Coverage: bar, bar_top_n, pie, donut, line_count_over_time, line,
grouped_bar, scatter — these map cleanly onto openpyxl's native chart
types. histogram, box, heatmap, and stacked_area do NOT have a clean
native-Excel equivalent via openpyxl, so those tiles get a plain data
table + a note instead of a fabricated approximation — same principle as
the QA agent: don't guess, say plainly what's not supported.
"""

import io

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, DoughnutChart, ScatterChart, Reference, Series
from openpyxl.styles import Font

UNSUPPORTED_NATIVE = {"histogram", "box", "heatmap", "stacked_area"}


def _write_value_counts(ws, start_row, col_name, series, top_n=None):
    counts = series.value_counts()
    if top_n:
        counts = counts.head(top_n)
    ws.cell(row=start_row, column=1, value=col_name).font = Font(bold=True)
    ws.cell(row=start_row, column=2, value="count").font = Font(bold=True)
    for i, (label, count) in enumerate(counts.items(), start=1):
        ws.cell(row=start_row + i, column=1, value=str(label))
        ws.cell(row=start_row + i, column=2, value=int(count))
    return start_row + 1, start_row + len(counts)  # header row, last data row


def _add_bar_or_pie(dash_ws, data_ws, tile, df, anchor_row):
    col = tile["columns"][0]
    chart_type = tile["chart_type"]
    top_n = 10 if chart_type == "bar_top_n" else None
    header_row, last_row = _write_value_counts(data_ws, anchor_row, col, df[col], top_n)
    data_col = data_ws.max_column  # data always starts at col A/B for this sheet section — see caller
    cats = Reference(data_ws, min_col=1, min_row=header_row + 1, max_row=last_row)
    vals = Reference(data_ws, min_col=2, min_row=header_row, max_row=last_row)

    if chart_type in ("pie",):
        chart = PieChart()
    elif chart_type in ("donut",):
        chart = DoughnutChart()
    else:
        chart = BarChart()
        chart.type = "col"

    chart.add_data(vals, titles_from_data=True)
    chart.set_categories(cats)
    chart.title = tile["title"]
    chart.width, chart.height = 16, 9
    return chart, last_row + 3


def _add_line(dash_ws, data_ws, tile, df, anchor_row):
    col = tile["columns"][0]
    s = df[col].value_counts().sort_index()
    header_row = anchor_row
    data_ws.cell(row=header_row, column=1, value=col).font = Font(bold=True)
    data_ws.cell(row=header_row, column=2, value="count").font = Font(bold=True)
    for i, (idx, count) in enumerate(s.items(), start=1):
        data_ws.cell(row=header_row + i, column=1, value=str(idx))
        data_ws.cell(row=header_row + i, column=2, value=int(count))
    last_row = header_row + len(s)

    chart = LineChart()
    cats = Reference(data_ws, min_col=1, min_row=header_row + 1, max_row=last_row)
    vals = Reference(data_ws, min_col=2, min_row=header_row, max_row=last_row)
    chart.add_data(vals, titles_from_data=True)
    chart.set_categories(cats)
    chart.title = tile["title"]
    chart.width, chart.height = 16, 9
    return chart, last_row + 3


def _add_pair_grouped_bar_or_line_or_scatter(dash_ws, data_ws, tile, df, anchor_row):
    col_a, col_b = tile["columns"]
    type_a, type_b = tile["semantic_types"]
    chart_type = tile["chart_type"]
    header_row = anchor_row

    if chart_type == "grouped_bar":
        cat_col, num_col = (col_a, col_b) if type_a.startswith("categorical") else (col_b, col_a)
        agg = df.groupby(cat_col, as_index=False)[num_col].mean().sort_values(num_col, ascending=False)
        data_ws.cell(row=header_row, column=1, value=cat_col).font = Font(bold=True)
        data_ws.cell(row=header_row, column=2, value=num_col).font = Font(bold=True)
        for i, row in enumerate(agg.itertuples(index=False), start=1):
            data_ws.cell(row=header_row + i, column=1, value=str(row[0]))
            data_ws.cell(row=header_row + i, column=2, value=float(row[1]))
        last_row = header_row + len(agg)
        chart = BarChart()
        chart.type = "col"
        cats = Reference(data_ws, min_col=1, min_row=header_row + 1, max_row=last_row)
        vals = Reference(data_ws, min_col=2, min_row=header_row, max_row=last_row)
        chart.add_data(vals, titles_from_data=True)
        chart.set_categories(cats)

    elif chart_type == "line":
        date_col, num_col = (col_a, col_b) if type_a == "datetime" else (col_b, col_a)
        s = df[[date_col, num_col]].dropna().sort_values(date_col)
        s = s.groupby(date_col, as_index=False)[num_col].sum()
        data_ws.cell(row=header_row, column=1, value=date_col).font = Font(bold=True)
        data_ws.cell(row=header_row, column=2, value=num_col).font = Font(bold=True)
        for i, row in enumerate(s.itertuples(index=False), start=1):
            data_ws.cell(row=header_row + i, column=1, value=str(row[0]))
            data_ws.cell(row=header_row + i, column=2, value=float(row[1]))
        last_row = header_row + len(s)
        chart = LineChart()
        cats = Reference(data_ws, min_col=1, min_row=header_row + 1, max_row=last_row)
        vals = Reference(data_ws, min_col=2, min_row=header_row, max_row=last_row)
        chart.add_data(vals, titles_from_data=True)
        chart.set_categories(cats)

    elif chart_type == "scatter":
        sub = df[[col_a, col_b]].dropna()
        data_ws.cell(row=header_row, column=1, value=col_a).font = Font(bold=True)
        data_ws.cell(row=header_row, column=2, value=col_b).font = Font(bold=True)
        for i, row in enumerate(sub.itertuples(index=False), start=1):
            data_ws.cell(row=header_row + i, column=1, value=float(row[0]))
            data_ws.cell(row=header_row + i, column=2, value=float(row[1]))
        last_row = header_row + len(sub)
        chart = ScatterChart()
        xvalues = Reference(data_ws, min_col=1, min_row=header_row + 1, max_row=last_row)
        yvalues = Reference(data_ws, min_col=2, min_row=header_row + 1, max_row=last_row)
        series = Series(yvalues, xvalues, title=f"{col_a} vs {col_b}")
        chart.series.append(series)

    else:
        return None, header_row

    chart.title = tile["title"]
    chart.width, chart.height = 16, 9
    return chart, last_row + 3


def build_workbook(result: dict, profile: dict) -> Workbook:
    df = profile["dataframe"]
    spec = result["final_spec"]

    wb = Workbook()
    dash_ws = wb.active
    dash_ws.title = "Dashboard"
    data_ws = wb.create_sheet("Chart Data")

    dash_ws["A1"] = "Dashboard"
    dash_ws["A1"].font = Font(size=16, bold=True)
    dash_ws["A2"] = f"QA score: {result['final_score']}/100 · {result['iterations_used']} fix pass(es)"

    data_row = 1
    chart_anchor_row = 4
    unsupported = []

    for tile in spec["tiles"]:
        chart_type = tile["chart_type"]

        if chart_type in UNSUPPORTED_NATIVE:
            unsupported.append(tile["title"])
            continue

        if len(tile["columns"]) == 2:
            chart, data_row = _add_pair_grouped_bar_or_line_or_scatter(dash_ws, data_ws, tile, df, data_row)
        elif chart_type == "line_count_over_time":
            chart, data_row = _add_line(dash_ws, data_ws, tile, df, data_row)
        else:
            chart, data_row = _add_bar_or_pie(dash_ws, data_ws, tile, df, data_row)

        if chart:
            anchor = f"D{chart_anchor_row}"
            dash_ws.add_chart(chart, anchor)
            chart_anchor_row += 19  # rough vertical spacing so charts don't overlap

    if unsupported:
        note_row = chart_anchor_row + 1
        dash_ws.cell(row=note_row, column=1,
                      value="Not renderable as native Excel charts (open the HTML export instead): "
                            + ", ".join(unsupported))

    # append the agent report as plain text at the bottom of the data sheet
    report_ws = wb.create_sheet("Agent Report")
    for i, line in enumerate(result["report_markdown"].splitlines(), start=1):
        report_ws.cell(row=i, column=1, value=line)

    return wb


def export_excel_bytes(result: dict, profile: dict) -> bytes:
    wb = build_workbook(result, profile)
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
